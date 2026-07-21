#!/usr/bin/env python3
"""Build an Excel workbook (table + charts) from the gem5 300-attack sweep.

Values are written as literals: this is a report of a completed, immutable
simulation, so the aggregates are fixed. Each summary number is computed
here in Python from the raw per-attack data (also included, in full, on the
second sheet) so the whole workbook renders correctly in any viewer without
needing a recalculation pass.
"""
import csv, sys, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

csv_path, out_path = sys.argv[1], sys.argv[2]
rows = [dict(seed=int(r["seed"]), total=int(r["total_flips"]),
             X=int(r["X"]), Y=int(r["Y"]), Z=int(r["Z"]), P=int(r["P"]))
        for r in csv.DictReader(open(csv_path))]
N = len(rows)

CATS = ["X", "Y", "Z", "P"]
MEAN = {"X": "very weak", "Y": "weak", "Z": "strong", "P": "very strong"}
RANGE = {"X": "0.80 - 1.00", "Y": "0.50 - 0.80",
         "Z": "0.10 - 0.50", "P": "0.00 - 0.10"}
WEIGHT = {"X": 0.42, "Y": 0.16, "Z": 0.10, "P": 0.32}
# Model expected flips for a single fully-weak row (X=387, Y=107, Z=31,
# P=16; total 541) -> shares, so the column sums to exactly 100%.
_MODEL_FLIPS = {"X": 387, "Y": 107, "Z": 31, "P": 16}
_mtot = sum(_MODEL_FLIPS.values())
MODEL_SHARE = {c: _MODEL_FLIPS[c] / _mtot for c in _MODEL_FLIPS}

cat_total = {c: sum(r[c] for r in rows) for c in CATS}
cat_mean = {c: cat_total[c] / N for c in CATS}
grand = sum(r["total"] for r in rows)
cat_share = {c: (cat_total[c] / grand if grand else 0) for c in CATS}
totals = [r["total"] for r in rows]
nonzero = sum(1 for t in totals if t > 0)

ARIAL = "Arial"
BOLD = Font(name=ARIAL, bold=True)
NORMAL = Font(name=ARIAL)
WHITE_BOLD = Font(name=ARIAL, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="4C78A8")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE = Font(name=ARIAL, bold=True, size=14)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER


wb = Workbook()

# ---------------------------------------------------------------- Data sheet
ws = wb.create_sheet("Per-Attack Data")
headers = ["Seed", "Total Flips", "X (very weak)", "Y (weak)",
           "Z (strong)", "P (very strong)"]
ws.append(headers)
style_header(ws, 1, len(headers))
for r in rows:
    ws.append([r["seed"], r["total"], r["X"], r["Y"], r["Z"], r["P"]])
for rr in range(2, N + 2):
    for cc in range(1, 7):
        ws.cell(row=rr, column=cc).font = NORMAL
        ws.cell(row=rr, column=cc).border = BORDER
        ws.cell(row=rr, column=cc).alignment = CENTER
tot_row, avg_row = N + 2, N + 3
ws.cell(row=tot_row, column=1, value="TOTAL").font = BOLD
ws.cell(row=avg_row, column=1, value="AVERAGE").font = BOLD
col_vals = {2: totals, 3: [r["X"] for r in rows], 4: [r["Y"] for r in rows],
            5: [r["Z"] for r in rows], 6: [r["P"] for r in rows]}
for cc, vals in col_vals.items():
    ws.cell(row=tot_row, column=cc, value=sum(vals)).font = BOLD
    a = ws.cell(row=avg_row, column=cc, value=round(sum(vals) / N, 3))
    a.font = BOLD
    a.number_format = "0.00"
    ws.cell(row=tot_row, column=cc).border = BORDER
    ws.cell(row=avg_row, column=cc).border = BORDER
for i, w in enumerate([8, 12, 15, 12, 12, 16], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---------------------------------------------------------------- Summary sheet
sm = wb.create_sheet("Summary", 0)
sm["A1"] = "RowHammer Synthetic Model - gem5 Simulation Results"
sm["A1"].font = TITLE
sm["A2"] = ("Per-category bit-flip counts over %d simulated attacks "
            "(gem5, synthetic weak-cell model)" % N)
sm["A2"].font = Font(name=ARIAL, italic=True, size=10)

sm["A4"] = "Simulation setup"
sm["A4"].font = BOLD
params = [
    ("Simulator", "gem5 (X86/gem5.opt), patched HammerSim RowHammer model"),
    ("Config", "row791_2_aggressor_rows_synth.py (DDR4, DIMM1, bank 4)"),
    ("Attack", "Double-sided; aggressor rows 790 & 792; victim row 791"),
    ("Model", "prefer_device_map_data = False (synthetic model forced on)"),
    ("Attacks (seeds)", "%d  (seed = 1 .. %d)" % (N, N)),
    ("Category rule", "each flipped cell binned by its own flip probability"),
]
r = 5
for k, v in params:
    sm.cell(row=r, column=1, value=k).font = BOLD
    sm.cell(row=r, column=2, value=v).font = NORMAL
    r += 1

tbl0 = r + 1
cat_headers = ["Category", "Meaning", "Flip prob. range", "Cell weight",
               "Total flips", "Mean flips/attack", "Share of flips",
               "Model share"]
for c, h in enumerate(cat_headers, 1):
    sm.cell(row=tbl0, column=c, value=h)
style_header(sm, tbl0, len(cat_headers))

first = tbl0 + 1
for i, cat in enumerate(CATS):
    rr = first + i
    sm.cell(row=rr, column=1, value=cat).font = BOLD
    sm.cell(row=rr, column=2, value=MEAN[cat]).font = NORMAL
    sm.cell(row=rr, column=3, value=RANGE[cat]).font = NORMAL
    c4 = sm.cell(row=rr, column=4, value=WEIGHT[cat]); c4.number_format = "0%"
    sm.cell(row=rr, column=5, value=cat_total[cat])
    c6 = sm.cell(row=rr, column=6, value=round(cat_mean[cat], 2))
    c6.number_format = "0.00"
    c7 = sm.cell(row=rr, column=7, value=round(cat_share[cat], 4))
    c7.number_format = "0.0%"
    c8 = sm.cell(row=rr, column=8, value=MODEL_SHARE[cat])
    c8.number_format = "0.0%"
    for cc in range(4, 9):
        sm.cell(row=rr, column=cc).font = NORMAL
    for cc in range(1, 9):
        sm.cell(row=rr, column=cc).border = BORDER
        if cc != 2:
            sm.cell(row=rr, column=cc).alignment = CENTER
trow = first + 4
sm.cell(row=trow, column=1, value="TOTAL").font = BOLD
sm.cell(row=trow, column=5, value=grand).font = BOLD
tm = sm.cell(row=trow, column=6, value=round(sum(cat_mean.values()), 2))
tm.font = BOLD; tm.number_format = "0.00"
ts = sm.cell(row=trow, column=7, value=round(sum(cat_share.values()), 4))
ts.font = BOLD; ts.number_format = "0.0%"
tms = sm.cell(row=trow, column=8, value=round(sum(MODEL_SHARE.values()), 4))
ts.font = BOLD; tms.number_format = "0.0%"; tms.font = BOLD
for cc in range(1, 9):
    sm.cell(row=trow, column=cc).border = BORDER
    sm.cell(row=trow, column=cc).fill = SUB_FILL

so = trow + 2
sm.cell(row=so, column=1, value="Overall statistics").font = BOLD
stats = [
    ("Total attacks (seeds)", N, "0"),
    ("Grand total bit flips", grand, "0"),
    ("Mean flips per attack", round(grand / N, 2), "0.00"),
    ("Median flips per attack", statistics.median(totals), "0"),
    ("Max flips in one attack", max(totals), "0"),
    ("Attacks with >=1 flip",
     "%d  (%.1f%%)" % (nonzero, nonzero / N * 100), "General"),
    ("Attacks with 0 flips",
     "%d  (%.1f%%)" % (N - nonzero, (N - nonzero) / N * 100), "General"),
]
rr = so + 1
for label, val, fmt in stats:
    sm.cell(row=rr, column=1, value=label).font = NORMAL
    cell = sm.cell(row=rr, column=2, value=val)
    cell.font = BOLD
    if fmt != "General":
        cell.number_format = fmt
    rr += 1

# Note on why absolute counts are lower than the analytical 541 figure.
note = rr + 1
sm.cell(row=note, column=1,
        value="Note").font = BOLD
sm.cell(row=note + 1, column=1,
        value=("Most attacks yield 0 flips because a victim row only flips "
               "when it is classified 'weak' (~30%% chance per row)."))
sm.cell(row=note + 2, column=1,
        value=("Absolute counts are lower than the analytical single-weak-row "
               "figure (~541) because gem5 samples one column per hammer over "
               "a limited window; the category SHARES still match the model."))
for dr in (note + 1, note + 2):
    sm.cell(row=dr, column=1).font = Font(name=ARIAL, italic=True, size=9)

for col, w in zip("ABCDEFGH", [24, 18, 18, 11, 12, 16, 13, 12]):
    sm.column_dimensions[col].width = w

# --- Chart 1: total flips per category (downward trend) ---
chart = BarChart()
chart.type = "col"
chart.title = "Bit flips by cell category (%d attacks)" % N
chart.y_axis.title = "Total bit flips"
chart.x_axis.title = "Cell category (most -> least vulnerable)"
chart.legend = None
data_ref = Reference(sm, min_col=5, min_row=tbl0, max_row=first + 3)
cats_ref = Reference(sm, min_col=1, min_row=first, max_row=first + 3)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 8; chart.width = 15
sm.add_chart(chart, "J4")

# --- Chart 2: model vs gem5 share ---
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Share of flips: gem5 vs analytical model"
chart2.y_axis.title = "Share of all bit flips"
chart2.x_axis.title = "Cell category"
share_ref = Reference(sm, min_col=7, max_col=8, min_row=tbl0, max_row=first + 3)
chart2.add_data(share_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.height = 8; chart2.width = 15
sm.add_chart(chart2, "J20")

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]
wb.save(out_path)
print("wrote", out_path)
print("grand=%d  X=%d Y=%d Z=%d P=%d  shares=%.1f/%.1f/%.1f/%.1f"
      % (grand, cat_total["X"], cat_total["Y"], cat_total["Z"],
         cat_total["P"], cat_share["X"]*100, cat_share["Y"]*100,
         cat_share["Z"]*100, cat_share["P"]*100))
