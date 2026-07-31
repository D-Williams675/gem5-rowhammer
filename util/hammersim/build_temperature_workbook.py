#!/usr/bin/env python3
"""Excel workbook for the temperature-dependent weak-cell (W0 + canary) model."""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

out_path, graph_png = sys.argv[1], sys.argv[2]

ARIAL = "Arial"
BOLD = Font(name=ARIAL, bold=True)
NORMAL = Font(name=ARIAL)
ITALIC = Font(name=ARIAL, italic=True, size=9)
WHITE_BOLD = Font(name=ARIAL, bold=True, color="FFFFFF")
HDR = PatternFill("solid", fgColor="4C78A8")
SUB = PatternFill("solid", fgColor="D9E1F2")
TITLE = Font(name=ARIAL, bold=True, size=14)
SEC = Font(name=ARIAL, bold=True, size=11.5)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center")

wb = Workbook()
ws = wb.active
ws.title = "Temperature Model"

ws["A1"] = "Temperature-Dependent Weak Cells in gem5  (weak set = W0 + canary(T))"
ws["A1"].font = TITLE
ws["A2"] = ("Models the canary-cell behavior from Orosa et al., \"SpyHammer\" "
            "(arXiv:2210.04084), Fig. 4")
ws["A2"].font = ITALIC

# ---------------- Setup ----------------
r = 4
ws.cell(row=r, column=1, value="Simulation setup").font = SEC
setup = [
    ("Simulator", "gem5 (X86/gem5.opt), patched HammerSim RowHammer model"),
    ("Config", "row791_temperature.py (DDR4, DIMM1, bank 4, victim row 791)"),
    ("Temperature range modeled", "50 - 95 C, in 5 C ranges (10 ranges)"),
    ("W0 (weak at all temperatures)", "30% of cells"),
    ("Canary (weak at exactly 1 range)", "2% of cells per temperature range"),
    ("Expected weak set at any temp.", "W0 + p% = 30% + 2% = 32%"),
]
r += 1
for k, v in setup:
    ws.cell(row=r, column=1, value=k).font = BOLD
    ws.cell(row=r, column=2, value=v).font = NORMAL
    r += 1

# ---------------- Cell counts ----------------
r += 1
ws.cell(row=r, column=1, value="Cell counts").font = SEC
r += 1
counts = [
    ("Columns per row (model column space)", 1024),
    ("Rows per bank (simulated DRAM)", 8192),
    ("Total cells per bank", 8388608),
    ("Rows checked in classification test", 200),
    ("TOTAL CELLS CHECKED (classification)", 204800),
    ("gem5 attack runs (60 seeds x 6 temps)", 360),
    ("Distinct cells that flipped in attacks", 965),
]
for k, v in counts:
    ws.cell(row=r, column=1, value=k).font = BOLD
    c = ws.cell(row=r, column=2, value=v)
    c.font = BOLD if "TOTAL" in k else NORMAL
    c.number_format = "#,##0"
    r += 1

# ---------------- Table A: exact classification ----------------
r += 1
ws.cell(row=r, column=1,
        value="TABLE A - Cell classification in the model (exact, "
              "204,800 cells x 10 temperature ranges)").font = SEC
r += 1
tblA0 = r
hdrA = ["Cell type", "Meaning", "Cells", "Share of all cells",
        "Target (model setting)"]
for c, h in enumerate(hdrA, 1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.font = WHITE_BOLD; cell.fill = HDR; cell.alignment = CENTER
    cell.border = BORDER
rowsA = [
    ("Never weak", "not weak at any temperature", 102277, 0.4994, None),
    ("Canary", "weak at exactly 1 temperature range", 41030, 0.2003, 0.20),
    ("W0", "weak at ALL temperature ranges", 61493, 0.3003, 0.30),
    ("Weak at 2-9 ranges", "would violate the disjoint rule", 0, 0.0, 0.0),
]
r += 1
for name, mean, cells, share, target in rowsA:
    ws.cell(row=r, column=1, value=name).font = BOLD
    ws.cell(row=r, column=2, value=mean).font = NORMAL
    c3 = ws.cell(row=r, column=3, value=cells); c3.number_format = "#,##0"
    c3.font = NORMAL
    c4 = ws.cell(row=r, column=4, value=share); c4.number_format = "0.00%"
    c4.font = NORMAL
    if target is not None:
        c5 = ws.cell(row=r, column=5, value=target)
        c5.number_format = "0%"; c5.font = NORMAL
    else:
        ws.cell(row=r, column=5, value="-").font = NORMAL
    for cc in range(1, 6):
        ws.cell(row=r, column=cc).border = BORDER
        if cc >= 3:
            ws.cell(row=r, column=cc).alignment = CENTER
    r += 1
# total row
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=2, value="all cells checked").font = NORMAL
tc = ws.cell(row=r, column=3, value=204800); tc.font = BOLD
tc.number_format = "#,##0"
tp = ws.cell(row=r, column=4, value=1.0); tp.font = BOLD
tp.number_format = "0.00%"
for cc in range(1, 6):
    ws.cell(row=r, column=cc).border = BORDER
    ws.cell(row=r, column=cc).fill = SUB
r += 1
ws.cell(row=r, column=1,
        value="Weak set at any single temperature = W0 + canary(that range) "
              "= 32.0% (verified at all 10 ranges).").font = ITALIC
r += 1
ws.cell(row=r, column=1,
        value="Zero cells are weak at 2-9 ranges, so canary sets are fully "
              "disjoint - the professor's requirement is met exactly.").font = ITALIC

# ---------------- Table B: observed flips ----------------
r += 3
ws.cell(row=r, column=1,
        value="TABLE B - Observed bit flips in gem5 attacks "
              "(60 seeds x 6 temperatures = 360 runs)").font = SEC
r += 1
hdrB = ["Temperature (C)", "Total bit flips",
        "Cells flipping ONLY at this temp (canary)"]
for c, h in enumerate(hdrB, 1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.font = WHITE_BOLD; cell.fill = HDR; cell.alignment = CENTER
    cell.border = BORDER
rowsB = [(50, 627, 82), (60, 633, 61), (70, 622, 45),
         (80, 637, 79), (90, 617, 67), (95, 626, 41)]
r += 1
for t, flips, canary in rowsB:
    ws.cell(row=r, column=1, value=t).font = NORMAL
    ws.cell(row=r, column=2, value=flips).font = NORMAL
    ws.cell(row=r, column=3, value=canary).font = NORMAL
    for cc in range(1, 4):
        ws.cell(row=r, column=cc).border = BORDER
        ws.cell(row=r, column=cc).alignment = CENTER
    r += 1
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=2, value=sum(x[1] for x in rowsB)).font = BOLD
ws.cell(row=r, column=3, value=sum(x[2] for x in rowsB)).font = BOLD
for cc in range(1, 4):
    ws.cell(row=r, column=cc).border = BORDER
    ws.cell(row=r, column=cc).fill = SUB
    ws.cell(row=r, column=cc).alignment = CENTER
r += 1
ws.cell(row=r, column=1,
        value="Total flips stay nearly constant across temperature "
              "(617-637): the AMOUNT of weakness is fixed, but WHICH cells "
              "are weak shifts with temperature.").font = ITALIC

# ---------------- Table C: cells by #temps flipped ----------------
r += 3
ws.cell(row=r, column=1,
        value="TABLE C - Cells by how many of the 6 tested temperatures "
              "they flipped at").font = SEC
r += 1
hdrC = ["# temperatures flipped at", "Cells", "Interpretation"]
for c, h in enumerate(hdrC, 1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.font = WHITE_BOLD; cell.fill = HDR; cell.alignment = CENTER
    cell.border = BORDER
rowsC = [
    ("1 of 6", 375, "canary-dominated (temperature-specific)"),
    ("2 of 6", 5, "non-weak cells flipping by chance"),
    ("3 of 6", 4, "non-weak cells flipping by chance"),
    ("4 of 6", 17, "non-weak cells flipping by chance"),
    ("5 of 6", 87, "W0 cells that missed one temperature by chance"),
    ("6 of 6", 477, "W0-dominated (weak at every temperature)"),
]
r += 1
for label, cells, interp in rowsC:
    ws.cell(row=r, column=1, value=label).font = NORMAL
    ws.cell(row=r, column=2, value=cells).font = NORMAL
    ws.cell(row=r, column=3, value=interp).font = NORMAL
    for cc in range(1, 4):
        ws.cell(row=r, column=cc).border = BORDER
    ws.cell(row=r, column=1).alignment = CENTER
    ws.cell(row=r, column=2).alignment = CENTER
    r += 1
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=2, value=965).font = BOLD
for cc in range(1, 4):
    ws.cell(row=r, column=cc).border = BORDER
    ws.cell(row=r, column=cc).fill = SUB
r += 1
ws.cell(row=r, column=1,
        value="IMPORTANT: being classified weak is NOT the same as flipping. "
              "Even non-weak cells keep a small (<=5%) flip chance, so the "
              "2-5 rows above are sampling noise, not model violations. "
              "The exact classification is Table A.").font = ITALIC

# ---------------- Embed the graph ----------------
img = XLImage(graph_png)
ratio = 980 / float(img.width)
img.width = 980
img.height = int(img.height * ratio)
ws.add_image(img, "G4")

for col, w in zip("ABCDEF", [36, 42, 14, 18, 20, 14]):
    ws.column_dimensions[col].width = w

wb.save(out_path)
print("wrote", out_path)
