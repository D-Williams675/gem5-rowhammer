#!/usr/bin/env python3
# Copyright (c) 2025 The Regents of the University of California
# All rights reserved.
#
# Builds an Excel workbook for the hardware aging results: a per-age
# table for each aging rate (flips, measured increase, predicted
# increase), a table of the worn-out cell counts |W(t)|, and the
# rendered figure embedded alongside.
#
# Usage:
#   util/hammersim/build_aging_workbook.py out.xlsx aging.csv graph.png

import argparse
import collections
import csv as _csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

ARIAL = "Arial"
BOLD = Font(name=ARIAL, bold=True)
NORMAL = Font(name=ARIAL)
ITALIC = Font(name=ARIAL, italic=True, size=9)
WHITE_BOLD = Font(name=ARIAL, bold=True, color="FFFFFF")
HDR = PatternFill("solid", fgColor="4C78A8")
SUB = PatternFill("solid", fgColor="D9E1F2")
TITLE = Font(name=ARIAL, bold=True, size=14)
SEC = Font(name=ARIAL, bold=True, size=11.5)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center")


def header(ws, row, labels):
    for c, h in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR
        cell.alignment = CENTER
        cell.border = BORDER


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("out")
    ap.add_argument("csv")
    ap.add_argument("png", nargs="?", default=None)
    ap.add_argument("--seeds", type=int, default=100)
    ap.add_argument("--bank-cells", type=int, default=8388608)
    args = ap.parse_args()

    rows = list(_csv.DictReader(open(args.csv)))
    by_rate = collections.defaultdict(list)
    for r in rows:
        by_rate[float(r["aging_rate"])].append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hardware Aging"

    ws["A1"] = "RowHammer Hardware Aging - gem5 Simulation Results"
    ws["A1"].font = TITLE
    ws["A2"] = ("As DRAM wears out, additional cells become RowHammer-weak. "
                "Weak set at time t = W0 + W(t).")
    ws["A2"].font = ITALIC

    r = 4
    ws.cell(row=r, column=1, value="Model and setup").font = SEC
    setup = [
        ("Aging model", "|W(t)| = aging_rate x t(weeks) x bank cells"),
        ("Cumulative?", "Yes - W(t1) is a subset of W(t2) for t1 < t2; a "
                        "worn-out cell never recovers"),
        ("Distribution", "W(t) cells spread randomly across the bank"),
        ("Bank size", args.bank_cells),
        ("Baseline weak set", "W0 = 30% + 2% canary = 32% at any one "
                              "temperature"),
        ("Attacks per data point", args.seeds),
        ("Simulator", "gem5 (X86/gem5.opt), patched HammerSim model"),
        ("Config", "row791_temperature.py, blast_radius 6, 50 C"),
    ]
    r += 1
    for k, v in setup:
        ws.cell(row=r, column=1, value=k).font = BOLD
        c = ws.cell(row=r, column=2, value=v)
        c.font = NORMAL
        if isinstance(v, int):
            c.number_format = "#,##0"
        r += 1

    # ---- one table per aging rate ----
    letter = ord("A")
    for rate in sorted(by_rate):
        r += 2
        ws.cell(row=r, column=1,
                value="TABLE %c - aging_rate = %g per week"
                      % (chr(letter), rate)).font = SEC
        letter += 1
        r += 1
        header(ws, r, ["Age (weeks)", "Age (years)", "Worn-out cells |W(t)|",
                       "Fraction of bank", "Bit flips",
                       "Measured increase", "Predicted increase"])
        r += 1
        for row in sorted(by_rate[rate], key=lambda x: float(x["age_weeks"])):
            w = float(row["age_weeks"])
            aged = rate * w * args.bank_cells
            vals = [w, w / 52.0, round(aged), rate * w, int(row["flips"]),
                    float(row["observed_increase"]),
                    float(row["predicted_increase"])]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = NORMAL
                cell.border = BORDER
                cell.alignment = CENTER
                if c == 2:
                    cell.number_format = "0.0"
                elif c == 3:
                    cell.number_format = "#,##0"
                elif c == 4:
                    cell.number_format = "0.0000%"
                elif c == 5:
                    cell.number_format = "#,##0"
                elif c in (6, 7):
                    cell.number_format = "+0.00%;-0.00%;0.00%"
            r += 1

    r += 1
    ws.cell(row=r, column=1,
            value="Predicted increase assumes only currently non-weak cells "
                  "can wear out: weak(t) = b + rate*t*(1-b) with b = 0.32, "
                  "so flips scale by weak(t)/b.").font = ITALIC
    r += 1
    ws.cell(row=r, column=1,
            value="At the default rate (1e-5/week) only ~0.5% of the bank "
                  "wears out over 10 years, so the effect is deliberately "
                  "small; the higher rate is shown to make the trend "
                  "legible.").font = ITALIC

    if args.png:
        img = XLImage(args.png)
        ratio = 1000 / float(img.width)
        img.width = 1000
        img.height = int(img.height * ratio)
        ws.add_image(img, "I4")

    for col, w in zip("ABCDEFGH", [24, 30, 22, 18, 14, 18, 18, 14]):
        ws.column_dimensions[col].width = w

    wb.save(args.out)
    print("wrote", args.out)


if __name__ == "__main__":
    main()
